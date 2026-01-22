#!/usr/bin/env python3
"""
MCP Server for Outlook Email, SharePoint, and Monday.com operations.
Provides CRUD operations for emails, files, and project management via Model Context Protocol.
"""
import sys
from datetime import datetime
from typing import Optional

from mcp.server.fastmcp import FastMCP

# Import clients
from download_emails import OutlookEmailDownloader, MICROSOFT_CONFIG
from monday_client import MondayClient, MONDAY_CONFIG

# Create MCP server
mcp = FastMCP("outlook-monday")

# Global client instances (authenticated lazily)
_outlook_client: Optional[OutlookEmailDownloader] = None
_monday_client: Optional[MondayClient] = None


def get_outlook_client() -> OutlookEmailDownloader:
    """Get authenticated Outlook email client (singleton)."""
    global _outlook_client
    if _outlook_client is None:
        _outlook_client = OutlookEmailDownloader()
        if not MICROSOFT_CONFIG.get("client_id"):
            raise RuntimeError("MS_CLIENT_ID not set. Check .env file.")
        if not _outlook_client.authenticate():
            raise RuntimeError("Authentication failed")
    return _outlook_client


def get_monday_client() -> MondayClient:
    """Get Monday.com client (singleton)."""
    global _monday_client
    if _monday_client is None:
        if not MONDAY_CONFIG.get("api_key"):
            raise RuntimeError("MONDAY_API_KEY not set. Check .env file.")
        _monday_client = MondayClient()
    return _monday_client


# Alias for backwards compatibility
def get_client() -> OutlookEmailDownloader:
    return get_outlook_client()


# =============================================================================
# CREATE
# =============================================================================

@mcp.tool()
def send_email(to: str, subject: str, body: str, html: bool = False) -> dict:
    """
    Send an email.

    Args:
        to: Recipient email address(es), comma-separated for multiple
        subject: Email subject line
        body: Email body content
        html: If True, send body as HTML. Default is plain text.

    Returns:
        dict with success status
    """
    try:
        client = get_client()
        recipients = [addr.strip() for addr in to.split(",")]
        body_type = "HTML" if html else "Text"
        success = client.send_email(recipients, subject, body, body_type)
        return {
            "success": success,
            "message": f"Email sent to {to}" if success else "Failed to send email"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# READ
# =============================================================================

@mcp.tool()
def list_emails_today(limit: int = 50) -> dict:
    """
    List emails received today.

    Args:
        limit: Maximum number of emails to return (default 50)

    Returns:
        dict with list of emails
    """
    try:
        client = get_client()
        emails = client.fetch_received_emails_today(download_attachments=False)
        emails = emails[:limit]
        return {
            "success": True,
            "count": len(emails),
            "emails": [
                {
                    "id": e.id,
                    "subject": e.subject,
                    "sender": e.sender,
                    "date": e.date,
                    "has_attachments": len(e.attachments) > 0,
                    "body_preview": e.body[:200] + "..." if len(e.body) > 200 else e.body
                }
                for e in emails
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def list_emails(start_date: str, end_date: str, limit: int = 50) -> dict:
    """
    List emails received within a date range.

    Args:
        start_date: Start date in YYYY-MM-DD format
        end_date: End date in YYYY-MM-DD format
        limit: Maximum number of emails to return (default 50)

    Returns:
        dict with list of emails
    """
    try:
        client = get_client()
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
        emails = client.fetch_emails_range(start, end, download_attachments=False)
        emails = emails[:limit]
        return {
            "success": True,
            "count": len(emails),
            "date_range": {"start": start_date, "end": end_date},
            "emails": [
                {
                    "id": e.id,
                    "subject": e.subject,
                    "sender": e.sender,
                    "date": e.date,
                    "has_attachments": len(e.attachments) > 0,
                    "body_preview": e.body[:200] + "..." if len(e.body) > 200 else e.body
                }
                for e in emails
            ]
        }
    except ValueError as e:
        return {"success": False, "error": f"Invalid date format. Use YYYY-MM-DD. {e}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def list_mail_folders() -> dict:
    """
    List all mail folders in the mailbox.

    Returns:
        dict with list of folders including Inbox, Archive, Sent Items, etc.
    """
    try:
        client = get_client()
        folders = client.list_mail_folders()
        return {
            "success": True,
            "count": len(folders),
            "folders": [
                {
                    "id": f.get("id", ""),
                    "name": f.get("displayName", ""),
                    "total_items": f.get("totalItemCount", 0),
                    "unread_items": f.get("unreadItemCount", 0)
                }
                for f in folders
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def list_emails_from_folder(
    folder: str = "Inbox",
    start_date: str = "",
    end_date: str = "",
    limit: int = 50
) -> dict:
    """
    List emails from a specific folder or all folders.

    Args:
        folder: Folder name - "Inbox", "Archive", "SentItems", "Drafts", "DeletedItems",
                or "All" to search all folders (default: Inbox)
        start_date: Optional start date filter (YYYY-MM-DD)
        end_date: Optional end date filter (YYYY-MM-DD)
        limit: Maximum number of emails to return (default 50)

    Returns:
        dict with list of emails from the specified folder
    """
    try:
        client = get_client()

        start = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
        end = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None

        emails = client.fetch_emails_from_folder(
            folder=folder,
            start_date=start,
            end_date=end,
            limit=limit,
            download_attachments=False
        )

        return {
            "success": True,
            "count": len(emails),
            "folder": folder,
            "emails": [
                {
                    "id": e.id,
                    "subject": e.subject,
                    "sender": e.sender,
                    "date": e.date,
                    "has_attachments": len(e.attachments) > 0,
                    "body_preview": e.body[:200] + "..." if len(e.body) > 200 else e.body
                }
                for e in emails
            ]
        }
    except ValueError as e:
        return {"success": False, "error": f"Invalid date format. Use YYYY-MM-DD. {e}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def search_emails(query: str, folder: str = "All", limit: int = 50) -> dict:
    """
    Search emails by keyword across subject, body, and sender.

    Args:
        query: Search query string
        folder: Folder to search - "All" (default), "Inbox", "Archive", "SentItems"
        limit: Maximum number of results (default 50)

    Returns:
        dict with list of matching emails
    """
    try:
        client = get_client()
        emails = client.search_emails(query=query, folder=folder, limit=limit)

        return {
            "success": True,
            "count": len(emails),
            "query": query,
            "folder": folder,
            "emails": [
                {
                    "id": e.id,
                    "subject": e.subject,
                    "sender": e.sender,
                    "date": e.date,
                    "has_attachments": len(e.attachments) > 0,
                    "body_preview": e.body[:200] + "..." if len(e.body) > 200 else e.body
                }
                for e in emails
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def get_email(email_id: str) -> dict:
    """
    Get a specific email by ID.

    Args:
        email_id: The email message ID

    Returns:
        dict with email details including full body
    """
    try:
        client = get_client()
        email = client.fetch_email_by_id(email_id, download_attachments=False)
        if email is None:
            return {"success": False, "error": "Email not found"}
        return {
            "success": True,
            "email": {
                "id": email.id,
                "subject": email.subject,
                "sender": email.sender,
                "recipients": email.recipients,
                "date": email.date,
                "body": email.body,
                "attachments": [
                    {"name": a.name, "size": a.size, "content_type": a.content_type}
                    for a in email.attachments
                ],
                "conversation_id": email.conversation_id
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def list_email_attachments(email_id: str) -> dict:
    """
    List all attachments for an email.

    Args:
        email_id: The email message ID

    Returns:
        dict with list of attachments including their IDs for downloading
    """
    try:
        client = get_client()
        attachments = client.list_attachments(email_id)
        return {
            "success": True,
            "count": len(attachments),
            "email_id": email_id,
            "attachments": attachments
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def download_attachment(email_id: str, attachment_id: str, save_to_disk: bool = True) -> dict:
    """
    Download an email attachment.

    Args:
        email_id: The email message ID
        attachment_id: The attachment ID (get from list_email_attachments)
        save_to_disk: If True, saves file to downloads folder and returns path.
                     If False, returns base64-encoded content.

    Returns:
        dict with attachment info and either file_path or content_base64
    """
    try:
        client = get_client()

        if save_to_disk:
            # Save to disk
            filepath = client.save_attachment(email_id, attachment_id)
            if not filepath:
                return {"success": False, "error": "Failed to download attachment"}
            return {
                "success": True,
                "message": f"Attachment saved to {filepath}",
                "file_path": str(filepath)
            }
        else:
            # Return base64 content
            attachment = client.download_attachment(email_id, attachment_id)
            if not attachment:
                return {"success": False, "error": "Failed to download attachment"}
            return {
                "success": True,
                "attachment": {
                    "name": attachment["name"],
                    "size": attachment["size"],
                    "content_type": attachment["content_type"],
                    "content_base64": attachment["content_base64"]
                }
            }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# UPDATE
# =============================================================================

@mcp.tool()
def mark_as_read(email_id: str, is_read: bool = True) -> dict:
    """
    Mark an email as read or unread.

    Args:
        email_id: The email message ID
        is_read: True to mark as read, False to mark as unread (default True)

    Returns:
        dict with success status
    """
    try:
        client = get_client()
        success = client.mark_email_read(email_id, is_read)
        status = "read" if is_read else "unread"
        return {
            "success": success,
            "message": f"Email marked as {status}" if success else f"Failed to mark email as {status}"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# DELETE
# =============================================================================

@mcp.tool()
def delete_email(email_id: str) -> dict:
    """
    Move an email to trash (soft delete).

    Args:
        email_id: The email message ID

    Returns:
        dict with success status
    """
    try:
        client = get_client()
        success = client.delete_email(email_id)
        return {
            "success": success,
            "message": "Email moved to trash" if success else "Failed to delete email"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# SHAREPOINT - READ
# =============================================================================

# Default site for convenience
DEFAULT_SITE = "xcellerateeq.sharepoint.com:/sites/DevTeam"


@mcp.tool()
def get_sharepoint_site(site_path: str = DEFAULT_SITE) -> dict:
    """
    Get SharePoint site information.

    Args:
        site_path: Site path like "xcellerateeq.sharepoint.com:/sites/DevTeam"

    Returns:
        dict with site info including site_id needed for other operations
    """
    try:
        client = get_client()
        site = client.get_sharepoint_site(site_path)
        if not site:
            return {"success": False, "error": "Site not found"}
        return {
            "success": True,
            "site": {
                "id": site.get("id", ""),
                "name": site.get("displayName", ""),
                "description": site.get("description", ""),
                "web_url": site.get("webUrl", ""),
                "created": site.get("createdDateTime", "")
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def list_sharepoint_drives(site_id: str) -> dict:
    """
    List document libraries (drives) in a SharePoint site.

    Args:
        site_id: The SharePoint site ID (get from get_sharepoint_site)

    Returns:
        dict with list of document libraries
    """
    try:
        client = get_client()
        drives = client.list_sharepoint_drives(site_id)
        return {
            "success": True,
            "count": len(drives),
            "drives": [
                {
                    "id": d.get("id", ""),
                    "name": d.get("name", ""),
                    "description": d.get("description", ""),
                    "web_url": d.get("webUrl", ""),
                    "drive_type": d.get("driveType", "")
                }
                for d in drives
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def list_sharepoint_files(site_id: str, drive_id: str = "", folder_path: str = "") -> dict:
    """
    List files in a SharePoint document library.

    Args:
        site_id: The SharePoint site ID
        drive_id: The drive/document library ID (optional, uses default if empty)
        folder_path: Path within the drive (optional, empty for root)

    Returns:
        dict with list of files
    """
    try:
        client = get_client()
        files = client.list_sharepoint_files(
            site_id,
            drive_id if drive_id else None,
            folder_path
        )
        return {
            "success": True,
            "count": len(files),
            "folder_path": folder_path or "/",
            "files": [
                {
                    "id": f.id,
                    "name": f.name,
                    "size": f.size,
                    "mime_type": f.mime_type,
                    "web_url": f.web_url,
                    "modified": f.modified,
                    "modified_by": f.modified_by
                }
                for f in files
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def search_sharepoint_files(site_id: str, query: str) -> dict:
    """
    Search for files in a SharePoint site.

    Args:
        site_id: The SharePoint site ID
        query: Search query string (searches file names and content)

    Returns:
        dict with list of matching files
    """
    try:
        client = get_client()
        files = client.search_sharepoint_files(site_id, query)
        return {
            "success": True,
            "count": len(files),
            "query": query,
            "files": [
                {
                    "id": f.id,
                    "name": f.name,
                    "size": f.size,
                    "mime_type": f.mime_type,
                    "web_url": f.web_url,
                    "modified": f.modified,
                    "parent_path": f.parent_path
                }
                for f in files
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def get_sharepoint_file_info(site_id: str, drive_id: str, item_id: str) -> dict:
    """
    Get detailed information about a SharePoint file.

    Args:
        site_id: The SharePoint site ID
        drive_id: The drive/document library ID
        item_id: The file item ID

    Returns:
        dict with file details
    """
    try:
        client = get_client()
        url = f"{client.graph_endpoint}/sites/{site_id}/drives/{drive_id}/items/{item_id}"
        response = __import__('requests').get(url, headers=client._get_headers())

        if not response.ok:
            return {"success": False, "error": f"File not found: {response.status_code}"}

        item = response.json()
        return {
            "success": True,
            "file": {
                "id": item.get("id", ""),
                "name": item.get("name", ""),
                "size": item.get("size", 0),
                "mime_type": item.get("file", {}).get("mimeType", ""),
                "web_url": item.get("webUrl", ""),
                "created": item.get("createdDateTime", ""),
                "modified": item.get("lastModifiedDateTime", ""),
                "created_by": item.get("createdBy", {}).get("user", {}).get("displayName", ""),
                "modified_by": item.get("lastModifiedBy", {}).get("user", {}).get("displayName", ""),
                "download_url": item.get("@microsoft.graph.downloadUrl", "")
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def download_sharepoint_file(
    site_id: str,
    drive_id: str,
    item_id: str,
    filename: str,
    output_dir: str = ""
) -> dict:
    """
    Download a SharePoint file and save it to disk.

    Args:
        site_id: The SharePoint site ID
        drive_id: The drive/document library ID
        item_id: The file item ID
        filename: Name to save the file as
        output_dir: Optional directory to save to (default: downloads/sharepoint)

    Returns:
        dict with file path or error
    """
    try:
        client = get_client()
        saved_path = client.save_sharepoint_file(
            site_id=site_id,
            drive_id=drive_id,
            item_id=item_id,
            filename=filename,
            output_dir=output_dir if output_dir else None
        )

        if saved_path:
            return {
                "success": True,
                "message": f"File saved successfully",
                "path": saved_path
            }
        else:
            return {"success": False, "error": "Failed to download file"}
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# MONDAY.COM - READ
# =============================================================================

@mcp.tool()
def list_monday_boards(limit: int = 50) -> dict:
    """
    List all Monday.com boards accessible to the user.

    Args:
        limit: Maximum number of boards to return (default 50)

    Returns:
        dict with list of boards
    """
    try:
        client = get_monday_client()
        boards = client.list_boards(limit=limit)
        return {
            "success": True,
            "count": len(boards),
            "boards": [
                {
                    "id": b.id,
                    "name": b.name,
                    "description": b.description,
                    "state": b.state,
                    "board_kind": b.board_kind
                }
                for b in boards
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def get_monday_board(board_id: str) -> dict:
    """
    Get details of a specific Monday.com board.

    Args:
        board_id: The board ID

    Returns:
        dict with board details including columns
    """
    try:
        client = get_monday_client()
        board = client.get_board(board_id)
        if not board:
            return {"success": False, "error": "Board not found"}

        columns = client.get_columns(board_id)
        groups = client.list_groups(board_id)

        return {
            "success": True,
            "board": {
                "id": board.id,
                "name": board.name,
                "description": board.description,
                "state": board.state,
                "board_kind": board.board_kind
            },
            "columns": [
                {"id": c.get("id"), "title": c.get("title"), "type": c.get("type")}
                for c in columns
            ],
            "groups": [
                {"id": g.id, "title": g.title, "color": g.color}
                for g in groups
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def list_monday_items(board_id: str, limit: int = 100) -> dict:
    """
    List items (rows) in a Monday.com board.

    Args:
        board_id: The board ID
        limit: Maximum number of items to return (default 100)

    Returns:
        dict with list of items
    """
    try:
        client = get_monday_client()
        items = client.list_items(board_id, limit=limit)
        return {
            "success": True,
            "count": len(items),
            "board_id": board_id,
            "items": [
                {
                    "id": item.id,
                    "name": item.name,
                    "group_id": item.group_id,
                    "state": item.state,
                    "created_at": item.created_at,
                    "updated_at": item.updated_at,
                    "column_values": {
                        k: v.get("text", "") for k, v in item.column_values.items()
                    }
                }
                for item in items
            ]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def get_monday_item(item_id: str) -> dict:
    """
    Get details of a specific Monday.com item.

    Args:
        item_id: The item ID

    Returns:
        dict with item details
    """
    try:
        client = get_monday_client()
        item = client.get_item(item_id)
        if not item:
            return {"success": False, "error": "Item not found"}

        return {
            "success": True,
            "item": {
                "id": item.id,
                "name": item.name,
                "board_id": item.board_id,
                "group_id": item.group_id,
                "state": item.state,
                "created_at": item.created_at,
                "updated_at": item.updated_at,
                "column_values": item.column_values
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# MONDAY.COM - WRITE
# =============================================================================

@mcp.tool()
def create_monday_item(board_id: str, item_name: str, group_id: str = "", column_values: str = "") -> dict:
    """
    Create a new item in a Monday.com board.

    Args:
        board_id: The board ID
        item_name: Name of the new item
        group_id: Optional group ID (uses first group if empty)
        column_values: Optional JSON string of column_id -> value mapping

    Returns:
        dict with created item details
    """
    try:
        import json
        client = get_monday_client()

        col_vals = None
        if column_values:
            col_vals = json.loads(column_values)

        item = client.create_item(
            board_id=board_id,
            item_name=item_name,
            group_id=group_id if group_id else None,
            column_values=col_vals
        )

        if not item:
            return {"success": False, "error": "Failed to create item"}

        return {
            "success": True,
            "message": f"Item '{item_name}' created",
            "item": {
                "id": item.id,
                "name": item.name,
                "board_id": item.board_id
            }
        }
    except json.JSONDecodeError as e:
        return {"success": False, "error": f"Invalid JSON for column_values: {e}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def update_monday_item(board_id: str, item_id: str, column_values: str) -> dict:
    """
    Update column values of a Monday.com item.

    Args:
        board_id: The board ID
        item_id: The item ID
        column_values: JSON string of column_id -> value mapping

    Returns:
        dict with success status
    """
    try:
        import json
        client = get_monday_client()

        col_vals = json.loads(column_values)
        success = client.update_item(board_id, item_id, col_vals)

        return {
            "success": success,
            "message": "Item updated" if success else "Failed to update item"
        }
    except json.JSONDecodeError as e:
        return {"success": False, "error": f"Invalid JSON for column_values: {e}"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def delete_monday_item(item_id: str) -> dict:
    """
    Delete a Monday.com item.

    Args:
        item_id: The item ID

    Returns:
        dict with success status
    """
    try:
        client = get_monday_client()
        success = client.delete_item(item_id)
        return {
            "success": success,
            "message": "Item deleted" if success else "Failed to delete item"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def move_monday_item(item_id: str, group_id: str) -> dict:
    """
    Move a Monday.com item to a different group.

    Args:
        item_id: The item ID
        group_id: Target group ID

    Returns:
        dict with success status
    """
    try:
        client = get_monday_client()
        success = client.move_item_to_group(item_id, group_id)
        return {
            "success": success,
            "message": f"Item moved to group {group_id}" if success else "Failed to move item"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# TASK ALIGNMENT - REMOTE FETCHERS
# =============================================================================

@mcp.tool()
def fetch_blueprint_context(task: str) -> dict:
    """
    Fetch blueprint architecture diagrams relevant to a task.

    Searches local Blueprint Diagrams folder for .mmd files matching task keywords.
    Use this before creating deliverables to understand the architecture.

    Args:
        task: Task description to find relevant blueprints for

    Returns:
        dict with relevant diagrams and their content summaries
    """
    try:
        from pathlib import Path

        blueprint_dir = Path("/Users/Mike/Library/Mobile Documents/com~apple~CloudDocs/2025-2030/Xenodex/greg/~non-code/outlook-fetcher/all_context/Blueprint Diagrams")

        if not blueprint_dir.exists():
            return {"success": False, "error": "Blueprint directory not found"}

        diagrams = []
        keywords = task.lower().split()

        for mmd_file in blueprint_dir.glob("*.mmd"):
            content = mmd_file.read_text()
            content_lower = content.lower()

            # Check if any keyword matches
            relevance_score = sum(1 for kw in keywords if kw in content_lower or kw in mmd_file.name.lower())

            if relevance_score > 0:
                # Extract key elements (subgraphs, nodes)
                lines = content.split('\n')
                elements = [l.strip() for l in lines if 'subgraph' in l.lower() or '[' in l][:10]

                diagrams.append({
                    "name": mmd_file.name,
                    "relevance_score": relevance_score,
                    "elements": elements,
                    "content_preview": content[:500] + "..." if len(content) > 500 else content
                })

        # Sort by relevance
        diagrams.sort(key=lambda x: x["relevance_score"], reverse=True)

        return {
            "success": True,
            "count": len(diagrams),
            "task": task,
            "diagrams": diagrams[:10]  # Top 10 most relevant
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def fetch_roadmap_context(task: str) -> dict:
    """
    Fetch roadmap entries relevant to a task from the v07 xlsx.

    Searches the Roadmap spreadsheet for features and components matching task keywords.
    Use this to find feature IDs, component IDs, owners, and status.

    Args:
        task: Task description to find relevant roadmap entries for

    Returns:
        dict with matching roadmap entries
    """
    try:
        from pathlib import Path
        import openpyxl

        roadmap_path = Path("/Users/Mike/Library/Mobile Documents/com~apple~CloudDocs/2025-2030/Xenodex/greg/~non-code/outlook-fetcher/downloads/sharepoint/Roadmap-SourceDoc-2026.01.17ss.v07.xlsx")

        if not roadmap_path.exists():
            return {"success": False, "error": "Roadmap file not found"}

        wb = openpyxl.load_workbook(roadmap_path, read_only=True, data_only=True)

        matches = []
        keywords = task.lower().split()

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = [cell.value for cell in sheet[1]] if sheet.max_row > 0 else []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_text = " ".join(str(cell or "") for cell in row).lower()

                # Check if any keyword matches
                if any(kw in row_text for kw in keywords):
                    row_dict = dict(zip(headers, row)) if headers else {"values": row}
                    matches.append({
                        "sheet": sheet_name,
                        "feature_id": row_dict.get("Feature ID") or row_dict.get("feature_id"),
                        "component_id": row_dict.get("Component ID") or row_dict.get("component_id"),
                        "name": row_dict.get("Component Name") or row_dict.get("Feature") or row_dict.get("Name"),
                        "status": row_dict.get("Status") or row_dict.get("status"),
                        "owner": row_dict.get("Owner") or row_dict.get("owner"),
                        "raw": {k: str(v)[:100] for k, v in row_dict.items() if v}
                    })

        wb.close()

        return {
            "success": True,
            "count": len(matches),
            "task": task,
            "matches": matches[:50]  # Limit to 50
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def fetch_requirements_context(task: str) -> dict:
    """
    Fetch success criteria and requirements relevant to a task.

    Searches deliverables CSVs for acceptance criteria and success metrics.
    Use this to understand what "done" means for a task.

    Args:
        task: Task description to find relevant requirements for

    Returns:
        dict with matching requirements and success criteria
    """
    try:
        from pathlib import Path
        import csv

        deliverables_dir = Path("/Users/Mike/Library/Mobile Documents/com~apple~CloudDocs/2025-2030/Xenodex/greg/~non-code/components/deliverables")

        if not deliverables_dir.exists():
            return {"success": False, "error": "Deliverables directory not found"}

        requirements = []
        keywords = task.lower().split()

        for csv_file in deliverables_dir.glob("**/*.csv"):
            try:
                with open(csv_file, 'r') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        row_text = " ".join(str(v or "") for v in row.values()).lower()

                        if any(kw in row_text for kw in keywords):
                            requirements.append({
                                "file": csv_file.name,
                                "component_id": row.get("Component ID"),
                                "component_name": row.get("Component Name"),
                                "success_criteria": row.get("Success Criteria"),
                                "acceptance_test": row.get("Acceptance Test"),
                                "done_means": row.get("Done Means")
                            })
            except Exception:
                continue

        return {
            "success": True,
            "count": len(requirements),
            "task": task,
            "requirements": requirements[:30]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def fetch_codebase_context(task: str, max_files: int = 20) -> dict:
    """
    Fetch codebase files relevant to a task from the xcellerate-eq repo.

    Searches the local codebase for Python files matching task keywords.
    Use this to understand existing implementations before creating new code.

    Args:
        task: Task description to find relevant code for
        max_files: Maximum number of files to return (default 20)

    Returns:
        dict with matching code files and their content previews
    """
    try:
        from pathlib import Path

        repo_path = Path("/Users/Mike/Library/Mobile Documents/com~apple~CloudDocs/2025-2030/Xenodex/greg/xcellerate-eq")

        if not repo_path.exists():
            return {"success": False, "error": "Codebase directory not found"}

        files = []
        keywords = task.lower().split()

        for py_file in repo_path.rglob("*.py"):
            if "__pycache__" in str(py_file) or ".git" in str(py_file):
                continue

            try:
                content = py_file.read_text()
                content_lower = content.lower()
                filename_lower = py_file.name.lower()

                # Check relevance
                relevance = sum(1 for kw in keywords if kw in content_lower or kw in filename_lower)

                if relevance > 0:
                    # Extract function/class names
                    import re
                    functions = re.findall(r'def (\w+)\(', content)[:5]
                    classes = re.findall(r'class (\w+)', content)[:3]

                    files.append({
                        "path": str(py_file.relative_to(repo_path)),
                        "relevance": relevance,
                        "functions": functions,
                        "classes": classes,
                        "preview": content[:300] + "..." if len(content) > 300 else content
                    })
            except Exception:
                continue

        # Sort by relevance
        files.sort(key=lambda x: x["relevance"], reverse=True)

        return {
            "success": True,
            "count": len(files),
            "task": task,
            "files": files[:max_files]
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def validate_output_alignment(
    output_description: str,
    output_content: str,
    checks: str = "blueprint,roadmap,requirements"
) -> dict:
    """
    Validate that an output (CSV, code, etc.) doesn't contradict sources of truth.

    Use this AFTER creating a deliverable to ensure it aligns with architecture,
    roadmap, and requirements. Returns conflicts that need to be resolved.

    Args:
        output_description: What the output is (e.g., "success criteria CSV for PLAT-F10")
        output_content: The actual content to validate (CSV text, code, etc.)
        checks: Comma-separated sources to check against (blueprint,roadmap,requirements)

    Returns:
        dict with validation results and any conflicts found
    """
    try:
        from pathlib import Path

        conflicts = []
        warnings = []
        matches = []

        output_lower = output_content.lower()
        checks_list = [c.strip() for c in checks.split(",")]

        # Check Blueprint alignment
        if "blueprint" in checks_list:
            blueprint_dir = Path("/Users/Mike/Library/Mobile Documents/com~apple~CloudDocs/2025-2030/Xenodex/greg/~non-code/outlook-fetcher/all_context/Blueprint Diagrams")

            if blueprint_dir.exists():
                # Check component_to_blueprint_mapping.csv
                mapping_file = blueprint_dir / "component_to_blueprint_mapping.csv"
                if mapping_file.exists():
                    import csv
                    with open(mapping_file, 'r') as f:
                        reader = csv.DictReader(f)
                        for row in reader:
                            comp_id = row.get("Component ID", "")
                            if comp_id and comp_id.lower() in output_lower:
                                if row.get("In Blueprint") == "Yes":
                                    matches.append(f"Component {comp_id} is in blueprint")
                                else:
                                    warnings.append(f"Component {comp_id} not documented in blueprint (gap)")

                # Check for threshold contradictions
                if "50%" in output_content and "dominance" in output_lower:
                    matches.append("Dominance threshold 50% matches blueprint")
                elif "40%" in output_content and "dominance" in output_lower:
                    conflicts.append("CONFLICT: Output uses 40% dominance threshold but blueprint specifies 50%")

                if "400ms" in output_content and "interruption" in output_lower:
                    matches.append("Interruption threshold 400ms matches blueprint")

        # Check Roadmap alignment
        if "roadmap" in checks_list:
            import openpyxl
            roadmap_path = Path("/Users/Mike/Library/Mobile Documents/com~apple~CloudDocs/2025-2030/Xenodex/greg/~non-code/outlook-fetcher/downloads/sharepoint/Roadmap-SourceDoc-2026.01.17ss.v07.xlsx")

            if roadmap_path.exists():
                # Extract component IDs from output
                import re
                component_ids = re.findall(r'[A-Z]+-F\d+-C\d+', output_content)
                feature_ids = re.findall(r'[A-Z]+-F\d+', output_content)

                if component_ids:
                    matches.append(f"Found {len(component_ids)} component IDs in output")

                # Verify IDs exist in roadmap
                wb = openpyxl.load_workbook(roadmap_path, read_only=True, data_only=True)
                roadmap_text = ""
                for sheet in wb.sheetnames:
                    for row in wb[sheet].iter_rows(values_only=True):
                        roadmap_text += " ".join(str(c or "") for c in row) + " "
                wb.close()

                for comp_id in component_ids:
                    if comp_id in roadmap_text:
                        matches.append(f"Component {comp_id} found in roadmap")
                    else:
                        warnings.append(f"Component {comp_id} not found in roadmap v07")

        # Check Requirements alignment
        if "requirements" in checks_list:
            deliverables_dir = Path("/Users/Mike/Library/Mobile Documents/com~apple~CloudDocs/2025-2030/Xenodex/greg/~non-code/components/deliverables")

            if deliverables_dir.exists():
                # Check success criteria format
                if "(<" in output_content or "(>" in output_content or "(±" in output_content:
                    matches.append("Success criteria follow <outcome> (<metric>) format")
                elif "success criteria" in output_lower:
                    warnings.append("Success criteria may not follow required format")

        # Determine overall status
        if conflicts:
            status = "CONFLICTS"
            emoji = "⚠"
        elif warnings and not matches:
            status = "WARNING"
            emoji = "?"
        else:
            status = "ALIGNED"
            emoji = "✓"

        return {
            "success": True,
            "status": status,
            "emoji": emoji,
            "output_description": output_description,
            "matches": matches,
            "conflicts": conflicts,
            "warnings": warnings,
            "recommendation": "Resolve conflicts before finalizing" if conflicts else "Output aligns with sources of truth"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def get_all_context(task: str) -> dict:
    """
    Fetch context from ALL sources of truth for a task in one call.

    Use this before starting work on a task to understand:
    - Blueprint: Architecture constraints
    - Roadmap: Feature/component IDs, owners, status
    - Requirements: Success criteria, acceptance tests
    - Codebase: Existing implementations

    Args:
        task: Task description to gather context for

    Returns:
        dict with combined context from all sources
    """
    try:
        blueprint = fetch_blueprint_context(task)
        roadmap = fetch_roadmap_context(task)
        requirements = fetch_requirements_context(task)
        codebase = fetch_codebase_context(task, max_files=10)

        return {
            "success": True,
            "task": task,
            "blueprint": {
                "diagrams_found": blueprint.get("count", 0),
                "relevant_diagrams": [d["name"] for d in blueprint.get("diagrams", [])[:5]]
            },
            "roadmap": {
                "matches_found": roadmap.get("count", 0),
                "features": list(set(m.get("feature_id") for m in roadmap.get("matches", []) if m.get("feature_id")))[:10]
            },
            "requirements": {
                "requirements_found": requirements.get("count", 0),
                "components": list(set(r.get("component_id") for r in requirements.get("requirements", []) if r.get("component_id")))[:10]
            },
            "codebase": {
                "files_found": codebase.get("count", 0),
                "relevant_files": [f["path"] for f in codebase.get("files", [])[:5]]
            },
            "summary": f"Found context in {sum([1 for x in [blueprint, roadmap, requirements, codebase] if x.get('count', 0) > 0])}/4 sources"
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# TASK ALIGNMENT - FULL CHECK
# =============================================================================

@mcp.tool()
def check_task_alignment(
    task: str,
    use_ai: bool = False,
    checks: str = "blueprint,codebase,roadmap,requirements,coherence",
    ai_model: str = "gpt-4o-mini"
) -> dict:
    """
    Check if a task aligns with all sources of truth (architecture, codebase, roadmap, requirements).

    Args:
        task: Task description to validate (e.g., "implement tone shift detector")
        use_ai: If True, use OpenAI for semantic comparison instead of keyword matching
        checks: Comma-separated list of checks to run (blueprint,codebase,roadmap,requirements,coherence)
        ai_model: OpenAI model to use when use_ai=True (default: gpt-4o-mini, or gpt-5-nano-2025-08-07)

    Returns:
        dict with alignment status and detailed report
    """
    try:
        from task_alignment import check_task_alignment as _check_alignment

        checks_list = [c.strip() for c in checks.split(",")]

        report = _check_alignment(
            task_description=task,
            checks=checks_list,
            use_ai=use_ai,
            ai_model=ai_model,
        )

        return {
            "success": True,
            "task": task,
            "overall_status": report.overall_status.value,
            "overall_status_emoji": report.overall_status.emoji,
            "checks": [
                {
                    "source": c.source,
                    "status": c.status.value,
                    "status_emoji": c.status.emoji,
                    "matches": c.matches,
                    "conflicts": c.conflicts,
                    "missing": c.missing,
                    "warnings": c.warnings,
                }
                for c in report.checks
            ],
            "markdown_report": report.to_markdown(),
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def quick_task_check(task: str) -> dict:
    """
    Quick task alignment check with default settings (keyword matching, all sources).

    Args:
        task: Task description to validate

    Returns:
        dict with alignment status summary
    """
    try:
        from task_alignment import check_task_alignment as _check_alignment

        report = _check_alignment(task_description=task)

        return {
            "success": True,
            "task": task,
            "status": report.overall_status.value,
            "emoji": report.overall_status.emoji,
            "summary": f"{report.overall_status.emoji} {report.overall_status.value.upper()}: {len([c for c in report.checks if c.matches])} sources with matches",
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


@mcp.tool()
def ai_task_check(task: str, model: str = "gpt-4o-mini") -> dict:
    """
    AI-powered semantic task alignment check using OpenAI.

    Understands synonyms and related concepts (e.g., "voice analysis" matches "tone detection").
    Requires OPENAI_API_KEY environment variable.

    Args:
        task: Task description to validate
        model: OpenAI model (gpt-4o-mini, gpt-4o, gpt-5-nano-2025-08-07)

    Returns:
        dict with AI-analyzed alignment report
    """
    try:
        from task_alignment import check_task_alignment as _check_alignment

        report = _check_alignment(
            task_description=task,
            use_ai=True,
            ai_model=model,
        )

        return {
            "success": True,
            "task": task,
            "model": model,
            "overall_status": report.overall_status.value,
            "overall_status_emoji": report.overall_status.emoji,
            "checks": [
                {
                    "source": c.source,
                    "status": c.status.value,
                    "matches": c.matches[:5],  # Limit for readability
                    "warnings": c.warnings[:3],
                }
                for c in report.checks
            ],
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("Starting Outlook/SharePoint/Monday MCP Server...", file=sys.stderr)
    mcp.run()
