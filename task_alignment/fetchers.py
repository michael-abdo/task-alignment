"""Context fetchers for Task Alignment System.

Each fetcher retrieves relevant context from a source of truth:
- Blueprint: Mermaid diagrams (Mike)
- Codebase: GitHub repositories (Youssef/Patrick)
- Roadmap: Monday.com + Roadmap xlsx (Shanti)
- Requirements: Email attachments (Greg)
"""

from __future__ import annotations

import os
import re
import subprocess
from pathlib import Path
from typing import List, Optional

from .models import (
    BlueprintContext,
    CodebaseContext,
    RoadmapContext,
    RequirementsContext,
    DiagramInfo,
    CodeFile,
    RoadmapMatch,
)


# Paths
BLUEPRINT_DIR = Path(__file__).parent.parent / "all_context" / "Blueprint Diagrams"
DOWNLOADS_DIR = Path(__file__).parent.parent / "downloads"
GREG_REPO_PATH = Path("/Users/Mike/Library/Mobile Documents/com~apple~CloudDocs/2025-2030/Xenodex/greg")


def extract_keywords(task: str) -> List[str]:
    """Extract searchable keywords from a task description."""
    stop_words = {
        "the", "a", "an", "and", "or", "but", "in", "on", "at", "to", "for",
        "of", "with", "by", "from", "is", "are", "was", "were", "be", "been",
        "being", "have", "has", "had", "do", "does", "did", "will", "would",
        "could", "should", "may", "might", "must", "shall", "can", "need",
        "implement", "create", "add", "update", "fix", "build", "make",
        "component", "feature", "system", "module", "function", "method",
    }

    words = re.findall(r'\b[a-zA-Z_][a-zA-Z0-9_]*\b', task.lower())
    keywords = [w for w in words if w not in stop_words and len(w) > 2]

    camel_case = re.findall(r'[A-Z][a-z]+(?:[A-Z][a-z]+)*', task)
    keywords.extend([w.lower() for w in camel_case])

    return list(set(keywords))


def parse_mermaid_components(content: str) -> List[str]:
    """Extract component names from Mermaid diagram content."""
    components = []

    node_pattern = r'\b([A-Z_][A-Z0-9_]*)\[|subgraph\s+"([^"]+)"'
    for match in re.finditer(node_pattern, content):
        if match.group(1):
            components.append(match.group(1))
        if match.group(2):
            components.append(match.group(2).replace(" ", "_"))

    return list(set(components))


def parse_mermaid_flows(content: str) -> List[str]:
    """Extract data flow descriptions from Mermaid diagram."""
    flows = []

    flow_pattern = r'(\w+)\s*(?:-->|->|-->>|-.->)\s*(\w+)'
    for match in re.finditer(flow_pattern, content):
        flows.append(f"{match.group(1)} -> {match.group(2)}")

    return flows


# ============================================================================
# BLUEPRINT FETCHER (Mike's Architecture)
# ============================================================================

def fetch_blueprint_context(task: str) -> BlueprintContext:
    """
    Fetch relevant architecture diagrams from local Blueprint Diagrams folder.

    Source: ~/outlook-fetcher/all_context/Blueprint Diagrams/
    """
    context = BlueprintContext()
    keywords = extract_keywords(task)

    if not BLUEPRINT_DIR.exists():
        return context

    for mmd_file in BLUEPRINT_DIR.glob("*.mmd"):
        content = mmd_file.read_text()

        relevance_score = sum(
            1 for kw in keywords
            if kw in content.lower() or kw in mmd_file.stem.lower()
        )

        if relevance_score > 0:
            components = parse_mermaid_components(content)
            flows = parse_mermaid_flows(content)

            diagram = DiagramInfo(
                filename=mmd_file.name,
                path=str(mmd_file),
                content=content,
                components=components,
                flows=flows,
            )
            context.add_diagram(diagram)

    return context


# ============================================================================
# CODEBASE FETCHER (Youssef/Patrick's Implementation)
# ============================================================================

def fetch_codebase_context(task: str, repo_path: Optional[Path] = None, max_files: int = 50) -> CodebaseContext:
    """
    Fetch relevant code from local repository or GitHub.

    Sources:
    - Local clone: xcellerate-eq
    - GitHub API (if configured)
    """
    context = CodebaseContext()
    keywords = extract_keywords(task)

    repo_path = repo_path or (GREG_REPO_PATH / "xcellerate-eq")

    if not repo_path.exists():
        return context

    files_checked = 0
    files_added = 0

    for pattern in ["**/*.py", "**/*.ts", "**/*.tsx", "**/*.js"]:
        if files_added >= max_files:
            break

        for code_file in repo_path.glob(pattern):
            if files_added >= max_files:
                break

            if any(skip in str(code_file) for skip in [
                "node_modules", "__pycache__", ".git", "dist", "build", ".next",
                "htmlcov", "coverage", ".pytest_cache", "venv", ".venv"
            ]):
                continue

            files_checked += 1
            if files_checked > 500:
                break

            try:
                content = code_file.read_text()
            except Exception:
                continue

            relevance_score = sum(
                1 for kw in keywords
                if kw in content.lower() or kw in code_file.stem.lower()
            )

            if relevance_score > 0:
                lang = "python" if code_file.suffix == ".py" else "typescript"

                functions = re.findall(r'(?:def|function|const)\s+(\w+)\s*\(', content)
                classes = re.findall(r'class\s+(\w+)', content)

                file_info = CodeFile(
                    path=str(code_file.relative_to(repo_path)),
                    content=content[:5000],
                    language=lang,
                    functions=functions,
                    classes=classes,
                )
                context.files.append(file_info)
                files_added += 1

                for kw in keywords:
                    if kw in [f.lower() for f in functions]:
                        context.existing_implementations.append(
                            f"{kw} in {code_file.name}"
                        )

    try:
        result = subprocess.run(
            ["pip", "list", "--format=freeze"],
            capture_output=True, text=True, timeout=5
        )
        if result.returncode == 0:
            context.available_dependencies = [
                line.split("==")[0].lower()
                for line in result.stdout.strip().split("\n")
                if line
            ]
    except Exception:
        pass

    return context


# ============================================================================
# ROADMAP FETCHER (Shanti's Project Management)
# ============================================================================

def fetch_roadmap_context(task: str) -> RoadmapContext:
    """
    Fetch roadmap data from Monday.com and local Roadmap files.

    Sources:
    - Monday.com boards
    - Roadmap xlsx files
    """
    import sys
    sys.path.insert(0, str(Path(__file__).parent.parent))

    context = RoadmapContext()
    keywords = extract_keywords(task)

    try:
        from monday_client import MondayClient

        client = MondayClient()
        if client.api_key:
            boards = client.list_boards(limit=10)

            for board in boards:
                items = client.list_items(board.id, limit=50)

                for item in items:
                    item_name_lower = item.name.lower()
                    relevance = sum(1 for kw in keywords if kw in item_name_lower)

                    if relevance > 0:
                        context.monday_items.append({
                            "id": item.id,
                            "name": item.name,
                            "board": board.name,
                            "status": item.column_values.get("status", {}).get("text", ""),
                            "relevance": relevance,
                        })

                        match = RoadmapMatch(
                            feature_id=f"MONDAY-{board.id}",
                            component_id=item.id,
                            component_name=item.name,
                            status=item.column_values.get("status", {}).get("text", "Unknown"),
                        )
                        context.matches.append(match)
    except Exception as e:
        context.monday_items.append({"error": f"Monday.com fetch failed: {e}"})

    roadmap_files = list(DOWNLOADS_DIR.glob("*Roadmap*.xlsx")) + \
                    list(DOWNLOADS_DIR.glob("*roadmap*.csv"))

    for roadmap_file in roadmap_files:
        try:
            if roadmap_file.suffix == ".xlsx":
                import openpyxl
                wb = openpyxl.load_workbook(roadmap_file, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(max_row=100, values_only=True):
                        row_text = " ".join(str(cell) for cell in row if cell)
                        relevance = sum(1 for kw in keywords if kw in row_text.lower())
                        if relevance > 0:
                            context.features[row_text[:100]] = {
                                "source": roadmap_file.name,
                                "sheet": sheet,
                            }
            elif roadmap_file.suffix == ".csv":
                import csv
                with open(roadmap_file, "r") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        row_text = " ".join(row)
                        relevance = sum(1 for kw in keywords if kw in row_text.lower())
                        if relevance > 0:
                            context.features[row_text[:100]] = {
                                "source": roadmap_file.name,
                            }
        except Exception:
            pass

    return context


# ============================================================================
# REQUIREMENTS FETCHER (Greg's Business Logic)
# ============================================================================

def fetch_requirements_context(task: str) -> RequirementsContext:
    """
    Fetch requirements from downloaded email attachments.

    Sources:
    - CSV files in downloads/
    - Excel files with acceptance criteria
    """
    context = RequirementsContext()
    keywords = extract_keywords(task)

    if not DOWNLOADS_DIR.exists():
        return context

    for csv_file in DOWNLOADS_DIR.glob("*.csv"):
        try:
            content = csv_file.read_text()

            relevance = sum(1 for kw in keywords if kw in content.lower())
            if relevance > 0:
                context.add_requirement_file(csv_file.name, content)

                if "acceptance" in csv_file.name.lower() or "acceptance" in content.lower():
                    import csv
                    from io import StringIO

                    reader = csv.DictReader(StringIO(content))
                    for row in reader:
                        row_text = " ".join(str(v) for v in row.values())
                        if any(kw in row_text.lower() for kw in keywords):
                            criteria = row.get("Acceptance Criteria") or row.get("acceptance_criteria")
                            if criteria:
                                context.acceptance_criteria.append(criteria)

                            success = row.get("Success Criteria") or row.get("success_criteria")
                            if success:
                                context.success_criteria.append(success)

                            hours = row.get("Est. Hours") or row.get("estimated_hours")
                            if hours:
                                try:
                                    context.estimated_hours = float(str(hours).replace("hrs", "").strip())
                                except ValueError:
                                    pass

                            loc = row.get("Est. LOC") or row.get("estimated_loc")
                            if loc:
                                try:
                                    context.estimated_loc = int(str(loc).replace("~", "").strip())
                                except ValueError:
                                    pass

        except Exception:
            pass

    for xlsx_file in DOWNLOADS_DIR.glob("*.xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_file, read_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                headers = []
                for row_idx, row in enumerate(ws.iter_rows(max_row=200, values_only=True)):
                    if row_idx == 0:
                        headers = [str(h).lower() if h else "" for h in row]
                        continue

                    row_text = " ".join(str(cell) for cell in row if cell)
                    relevance = sum(1 for kw in keywords if kw in row_text.lower())

                    if relevance > 0:
                        row_dict = dict(zip(headers, row))

                        for key in ["acceptance criteria", "acceptance_criteria", "acc_criteria"]:
                            if key in row_dict and row_dict[key]:
                                context.acceptance_criteria.append(str(row_dict[key]))

                        for key in ["success criteria", "success_criteria", "success"]:
                            if key in row_dict and row_dict[key]:
                                context.success_criteria.append(str(row_dict[key]))

        except Exception:
            pass

    return context
